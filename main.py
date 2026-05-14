import json

import re
from pathlib import Path
from typing import Any, Dict, List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter


REMOVE_COLUMNS = [
    "medianResTime",
    "throughput",
    "receivedKBytesPerSec",
    "sentKBytesPerSec",
]

TIME_COLUMNS_MS_TO_SEC = {
    "meanResTime": "Avg ResTime in sec",
    "minResTime": "Min ResTime in sec",
    "maxResTime": "MaxRes Time in sec",
    "pct1ResTime": "90thPercentile Resp Time in Sec",
    "pct2ResTime": "95thPercentile Resp Time in Sec",
    "pct3ResTime": "99thPercentile Resp Time in Sec",
}


def is_transaction(name: str) -> bool:
    return bool(re.match(r"^T\d{2}", str(name).strip()))


def split_api_name(name: str) -> Tuple[str, str, str]:
    parts = str(name).split("/")
    if len(parts) >= 3:
        return parts[0], parts[1], "/".join(parts[2:])
    if len(parts) == 2:
        return parts[0], parts[1], ""
    return str(name), "", ""


def load_statistics_json(json_path: str | Path) -> pd.DataFrame:
    with open(json_path, "r", encoding="utf-8-sig") as file:
        data: Dict[str, Dict[str, Any]] = json.load(file)

    rows: List[Dict[str, Any]] = []
    for key, value in data.items():
        row = dict(value)
        row["transaction"] = row.get("transaction", key)
        rows.append(row)

    return pd.DataFrame(rows)



def parse_report_metadata(json_path: str | Path) -> Dict[str, str]:
    """
    Extract report context from the uploaded JSON filename.

    Supported examples:
    - CiscoIQ-SaaS-Support-Services_100Users_100KDevices_APJC_1Hour_April-19-2026_Report.json
    - CiscoIQ-SaaS-Support-Services_50Users_20KDevices_APJC_April15_Report.json
    - file names containing 10/13 digit epoch timestamps
    """
    from datetime import datetime, timezone

    filename = Path(json_path).name
    name = Path(json_path).stem

    def find_or_na(pattern: str, group: int = 1, flags: int = re.IGNORECASE) -> str:
        match = re.search(pattern, name, flags)
        return match.group(group) if match else "N/A"

    users = find_or_na(r"(\d+)\s*[_\-\s]*Users?")
    if users != "N/A":
        users = f"{users} Users"

    devices = find_or_na(r"(\d+(?:\.\d+)?\s*K?)\s*[_\-\s]*Devices?")
    if devices != "N/A":
        devices = f"{devices.replace(' ', '')} Devices"

    known_regions = ["APJC", "AMER", "EMEA", "LATAM", "NA", "EU", "US", "INDIA"]
    region = "N/A"
    upper_name = name.upper()
    for candidate in known_regions:
        if re.search(rf"(?:^|[_\-\s]){candidate}(?:$|[_\-\s])", upper_name):
            region = candidate
            break

    # Duration: 1_Hour, 2_Hour, 1Hour, 90_Min, 30Minutes, etc.
    duration = "N/A"
    duration_match = re.search(
        r"(\d+(?:\.\d+)?)\s*[_\-\s]*(hours?|hrs?|hr|minutes?|mins?|min)(?=$|[_\-\s])",
        name,
        re.IGNORECASE,
    )
    if duration_match:
        value = duration_match.group(1)
        unit = duration_match.group(2).lower()
        if unit.startswith("hour") or unit.startswith("hr"):
            duration = f"{value} Hour" if value == "1" else f"{value} Hours"
        else:
            duration = f"{value} Minute" if value == "1" else f"{value} Minutes"

    # Date from Month-Day-Year / MonthDayYear / Month-Day / MonthDay.
    month_pattern = (
        r"(Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|"
        r"Jul(?:y)?|Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)"
    )
    date = "N/A"
    date_match = re.search(
        month_pattern + r"[_\-\s]*(\d{1,2})(?:[_\-\s]*(\d{2,4}))?",
        name,
        re.IGNORECASE,
    )
    if date_match:
        month = date_match.group(1)
        day = date_match.group(2)
        year = date_match.group(3)
        if year and len(year) == 2:
            year = "20" + year
        date = f"{month} {int(day)}" + (f", {year}" if year else "")

    # Fallback: ISO-like date.
    if date == "N/A":
        iso_match = re.search(r"(20\d{2})[_\-\s]*(\d{1,2})[_\-\s]*(\d{1,2})", name)
        if iso_match:
            year, month, day = iso_match.groups()
            date = f"{year}-{int(month):02d}-{int(day):02d}"

    # Fallback: epoch timestamp, 10 or 13 digit.
    if date == "N/A":
        for epoch_candidate in re.findall(r"(?<!\d)(\d{10}|\d{13})(?!\d)", name):
            try:
                epoch_value = int(epoch_candidate)
                if len(epoch_candidate) == 13:
                    epoch_value = epoch_value / 1000
                dt = datetime.fromtimestamp(epoch_value, tz=timezone.utc)
                # Avoid accidental tiny/invalid timestamps.
                if 2000 <= dt.year <= 2100:
                    date = dt.strftime("%b %d, %Y")
                    break
            except Exception:
                pass

    return {
        "Report File": filename,
        "Concurrent Users": users,
        "Devices Count": devices,
        "Date": date,
        "Duration": duration,
        "Region": region,
    }


def apply_common_column_cleanup(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    for col in REMOVE_COLUMNS:
        if col in df.columns:
            df = df.drop(columns=[col])

    for old_col, new_col in TIME_COLUMNS_MS_TO_SEC.items():
        if old_col in df.columns:
            df[old_col] = pd.to_numeric(df[old_col], errors="coerce") / 1000
            df = df.rename(columns={old_col: new_col})

    return df


def add_api_sla_columns(apis_df: pd.DataFrame) -> pd.DataFrame:
    apis_df = apis_df.copy()
    apis_df["SLA Sec"] = apis_df["Feature"].astype(str).str.upper().str.startswith("ASKAI").map({True: 10, False: 2})
    apis_df["SLA Rule"] = apis_df["SLA Sec"].map(
        lambda x: "AskAI APIs SLA < 10 sec" if x == 10 else "Assets, Assessments, Home, Settings and Support APIs SLA < 2 sec"
    )
    def row_pass(row: pd.Series) -> str:
        threshold = float(row["SLA Sec"])
        avg_v = float(pd.to_numeric(row.get("Avg ResTime in sec"), errors="coerce") or 0)
        min_v = float(pd.to_numeric(row.get("Min ResTime in sec"), errors="coerce") or 0)
        max_v = float(pd.to_numeric(row.get("MaxRes Time in sec"), errors="coerce") or 0)
        p95_v = float(pd.to_numeric(row.get("95thPercentile Resp Time in Sec"), errors="coerce") or 0)
        return "PASS" if (avg_v <= threshold and min_v <= threshold and max_v <= threshold and p95_v <= threshold) else "FAIL"

    apis_df["SLA Status"] = apis_df.apply(row_pass, axis=1)
    apis_df["SLA Breach Sec"] = apis_df.apply(
        lambda row: max((pd.to_numeric(row.get("Avg ResTime in sec"), errors="coerce") or 0) - row["SLA Sec"], 0),
        axis=1,
    )
    return apis_df


def order_columns(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    if sheet_name == "APIs":
        # Do not show SLA helper columns in APIs sheet.
        # SLA calculations are still used for response-time cell highlighting.
        preferred = [
            "Feature",
            "Scenario",
            "Endpoint",
            "sampleCount",
            "errorCount",
            "errorPct",
            "Avg ResTime in sec",
            "Min ResTime in sec",
            "MaxRes Time in sec",
            "90thPercentile Resp Time in Sec",
            "95thPercentile Resp Time in Sec",
            "99thPercentile Resp Time in Sec",
        ]
        hidden = {"transaction", "SLA Sec", "SLA Rule", "SLA Status", "SLA Breach Sec"}
        remaining = [c for c in df.columns if c not in preferred and c not in hidden]
        return df[[c for c in preferred if c in df.columns] + remaining]

    preferred = [
        "transaction",
        "sampleCount",
        "errorCount",
        "errorPct",
        "Avg ResTime in sec",
        "Min ResTime in sec",
        "MaxRes Time in sec",
        "90thPercentile Resp Time in Sec",
        "95thPercentile Resp Time in Sec",
        "99thPercentile Resp Time in Sec",
    ]
    remaining = [c for c in df.columns if c not in preferred]
    return df[[c for c in preferred if c in df.columns] + remaining]


def build_single_report_frames(json_path: str | Path):
    df = load_statistics_json(json_path)

    transactions_df = df[df["transaction"].apply(is_transaction)].copy()
    apis_df = df[~df["transaction"].apply(is_transaction)].copy()
    errors_df = df[pd.to_numeric(df.get("errorCount", 0), errors="coerce").fillna(0) > 0].copy()

    split_df = apis_df["transaction"].apply(lambda x: pd.Series(split_api_name(x)))
    split_df.columns = ["Feature", "Scenario", "Endpoint"]
    apis_df = pd.concat([split_df, apis_df], axis=1)

    transactions_df = order_columns(apply_common_column_cleanup(transactions_df), "Transactions")
    errors_df = order_columns(apply_common_column_cleanup(errors_df), "Errors")
    apis_df = order_columns(add_api_sla_columns(apply_common_column_cleanup(apis_df)), "APIs")

    return {
        "Transactions": transactions_df,
        "Errors": errors_df,
        "APIs": apis_df,
        "Run_Info": pd.DataFrame([parse_report_metadata(json_path)]),
    }


def bucket_headers_for_track(track: str) -> List[str]:
    if str(track).upper().startswith("ASKAI"):
        return ["0 - 10s in %", "10 - 20s in %", "20 - 30s in %", "> 30s in %"]
    return ["0 - 2s in %", "3 - 4s in %", "4 - 6s in %", "> 6s in %"]


def bucket_index(seconds: float, is_askai: bool) -> int:
    value = float(seconds)
    if is_askai:
        if value <= 10:
            return 0
        if value <= 20:
            return 1
        if value <= 30:
            return 2
        return 3

    if value <= 2:
        return 0
    if value <= 4:
        return 1
    if value <= 6:
        return 2
    return 3


def prepare_api_df_for_track(json_path: str | Path) -> pd.DataFrame:
    df = load_statistics_json(json_path)
    df = df[~df["transaction"].apply(is_transaction)].copy()
    split_df = df["transaction"].apply(lambda x: pd.Series(split_api_name(x)))
    split_df.columns = ["Feature", "Scenario", "Endpoint"]
    df = pd.concat([split_df, df], axis=1)

    df["avg_sec"] = pd.to_numeric(df["meanResTime"], errors="coerce") / 1000
    df["min_sec"] = pd.to_numeric(df["minResTime"], errors="coerce") / 1000
    df["max_sec"] = pd.to_numeric(df["maxResTime"], errors="coerce") / 1000
    return df


def track_metric_values(df: pd.DataFrame, track: str, metric: str) -> List[Any]:
    g = df[df["Feature"] == track].copy()
    if g.empty:
        return ["", "", "", "", ""]

    metric_to_col = {
        "Avg": "avg_sec",
        "Min": "min_sec",
        "Max": "max_sec",
    }
    col = metric_to_col[metric]
    is_askai = str(track).upper().startswith("ASKAI")

    total_apis = len(g)
    counts = [0, 0, 0, 0]

    for value in pd.to_numeric(g[col], errors="coerce").dropna():
        counts[bucket_index(value, is_askai)] += 1

    percentages = [round((count / total_apis) * 100, 2) for count in counts]
    max_seconds = round(float(pd.to_numeric(g[col], errors="coerce").max()), 2)
    return percentages + [max_seconds]









def build_track_comparison_matrix(json_paths: List[str | Path], labels: List[str]) -> List[List[Any]]:
    prepared = [prepare_api_df_for_track(path) for path in json_paths]
    all_tracks = sorted(
        track
        for track in set().union(*[set(df["Feature"].dropna().astype(str)) for df in prepared])
        if track
        and track.strip().lower() != "total"
        and "select customer" not in track.strip().lower()
    )

    askai_tracks = [track for track in all_tracks if str(track).upper().startswith("ASKAI")]
    other_tracks = [track for track in all_tracks if not str(track).upper().startswith("ASKAI")]

    report_title = " vs ".join(labels)

    def metric_values_for_tracks(df: pd.DataFrame, tracks: List[str], metric: str) -> List[Any]:
        g = df[df["Feature"].isin(tracks)].copy()
        if g.empty:
            return ["", "", "", "", ""]

        metric_to_col = {
            "Avg": "avg_sec",
            "Min": "min_sec",
            "Max": "max_sec",
        }
        col = metric_to_col[metric]

        is_askai_section = all(str(track).upper().startswith("ASKAI") for track in tracks)
        values = pd.to_numeric(g[col], errors="coerce").dropna()
        if values.empty:
            return ["", "", "", "", ""]

        counts = [0, 0, 0, 0]
        for value in values:
            counts[bucket_index(value, is_askai_section)] += 1

        percentages = [round((count / len(values)) * 100, 2) for count in counts]
        max_seconds = round(float(values.max()), 2)
        return percentages + [max_seconds]

    def add_section(matrix: List[List[Any]], title: str, tracks: List[str], bucket_headers: List[str]) -> None:
        if not tracks:
            return

        section_width = 2 + (len(labels) * 6)

        matrix.append([f"{title} - {report_title}"] + [""] * (section_width - 1))

        header_row = ["Track", "Metric"]
        for _ in labels:
            header_row += bucket_headers + ["Max Seconds", ""]
        matrix.append(header_row)

        # Total rows are placed immediately below the header so they are easy to find.
        for metric in ["Avg", "Min", "Max"]:
            row = ["Total" if metric == "Avg" else "", metric]
            for df in prepared:
                row += metric_values_for_tracks(df, tracks, metric)
                row += [""]
            matrix.append(row)

        for track in tracks:
            for metric in ["Avg", "Min", "Max"]:
                row = [track if metric == "Avg" else "", metric]
                for df in prepared:
                    row += track_metric_values(df, track, metric)
                    row += [""]
                matrix.append(row)

        matrix.append([""] * section_width)

    matrix: List[List[Any]] = []

    add_section(
        matrix,
        "AskAI Tracks",
        askai_tracks,
        ["0-10sec %", "10-20sec %", "20-30sec %", ">30sec %"],
    )

    add_section(
        matrix,
        "Assets / Assessments / Home / Settings / Support Tracks",
        other_tracks,
        ["0-2sec %", "3-4sec %", "4-6sec %", ">6sec %"],
    )

    return matrix


def prepare_compare_frame(df: pd.DataFrame, label: str) -> pd.DataFrame:
    base = df.copy()
    base["transaction"] = base["transaction"].astype(str)
    needed = [
        "transaction",
        "sampleCount",
        "errorCount",
        "errorPct",
        "meanResTime",
        "pct1ResTime",
        "pct2ResTime",
        "pct3ResTime",
    ]
    available = [c for c in needed if c in base.columns]
    base = base[available].copy()

    for col in ["meanResTime", "pct1ResTime", "pct2ResTime", "pct3ResTime"]:
        if col in base.columns:
            base[col] = pd.to_numeric(base[col], errors="coerce") / 1000

    rename_map = {
        "sampleCount": f"{label} Sample Count",
        "errorCount": f"{label} Error Count",
        "errorPct": f"{label} Error %",
        "meanResTime": f"{label} Avg Sec",
        "pct1ResTime": f"{label} 90th Sec",
        "pct2ResTime": f"{label} 95th Sec",
        "pct3ResTime": f"{label} 99th Sec",
    }
    return base.rename(columns=rename_map)


def build_comparison(json_paths: List[str | Path], labels: List[str]) -> pd.DataFrame:
    compare_df = prepare_compare_frame(load_statistics_json(json_paths[0]), labels[0])

    for path, label in zip(json_paths[1:], labels[1:]):
        next_df = prepare_compare_frame(load_statistics_json(path), label)
        compare_df = compare_df.merge(next_df, on="transaction", how="outer")

    baseline = labels[0]
    latest = labels[-1]

    if f"{baseline} Avg Sec" in compare_df.columns and f"{latest} Avg Sec" in compare_df.columns:
        compare_df["Avg Sec Diff"] = compare_df[f"{latest} Avg Sec"] - compare_df[f"{baseline} Avg Sec"]
        compare_df["Avg Sec Diff %"] = (compare_df["Avg Sec Diff"] / compare_df[f"{baseline} Avg Sec"]) * 100

    if f"{baseline} 90th Sec" in compare_df.columns and f"{latest} 90th Sec" in compare_df.columns:
        compare_df["90th Sec Diff"] = compare_df[f"{latest} 90th Sec"] - compare_df[f"{baseline} 90th Sec"]
        compare_df["90th Sec Diff %"] = (compare_df["90th Sec Diff"] / compare_df[f"{baseline} 90th Sec"]) * 100

    if f"{baseline} Error Count" in compare_df.columns and f"{latest} Error Count" in compare_df.columns:
        compare_df["Error Count Diff"] = compare_df[f"{latest} Error Count"] - compare_df[f"{baseline} Error Count"]

    split_df = compare_df["transaction"].apply(lambda x: pd.Series(split_api_name(x)))
    split_df.columns = ["Feature", "Scenario", "Endpoint"]
    compare_df = pd.concat([split_df, compare_df], axis=1)

    first_cols = ["Feature", "Scenario", "Endpoint", "transaction"]
    other_cols = [c for c in compare_df.columns if c not in first_cols]
    return compare_df[first_cols + other_cols]


def build_report(json_path: str | Path, output_excel_path: str | Path) -> None:
    frames = build_single_report_frames(json_path)
    track_matrix = build_track_comparison_matrix([json_path], [Path(json_path).stem])
    write_excel(frames, output_excel_path, track_matrix=track_matrix)



def build_apis_comparison(json_paths: List[str | Path], labels: List[str]) -> pd.DataFrame:
    """
    Build API-level side-by-side comparison.
    Each uploaded report gets its own Feature, Scenario, Endpoint columns
    plus selected metric columns.
    """
    comparison_df = None

    metric_cols = [
        "sampleCount",
        "errorCount",
        "errorPct",
        "meanResTime",
        "minResTime",
        "maxResTime",
        "pct1ResTime",
        "pct2ResTime",
        "pct3ResTime",
    ]

    metric_rename = {
        "sampleCount": "Sample Count",
        "errorCount": "Error Count",
        "errorPct": "Error %",
        "meanResTime": "Avg ResTime in sec",
        "minResTime": "Min ResTime in sec",
        "maxResTime": "MaxRes Time in sec",
        "pct1ResTime": "90thPercentile Resp Time in Sec",
        "pct2ResTime": "95thPercentile Resp Time in Sec",
        "pct3ResTime": "99thPercentile Resp Time in Sec",
    }

    for path, label in zip(json_paths, labels):
        df = load_statistics_json(path)
        df = df[~df["transaction"].apply(is_transaction)].copy()

        split_df = df["transaction"].apply(lambda x: pd.Series(split_api_name(x)))
        split_df.columns = [f"{label} Feature", f"{label} Scenario", f"{label} Endpoint"]
        df = pd.concat([split_df, df], axis=1)

        # Convert response-time metrics from ms to sec.
        for col in ["meanResTime", "minResTime", "maxResTime", "pct1ResTime", "pct2ResTime", "pct3ResTime"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce") / 1000

        keep_cols = ["transaction", f"{label} Feature", f"{label} Scenario", f"{label} Endpoint"]
        keep_cols += [c for c in metric_cols if c in df.columns]
        run_df = df[keep_cols].copy()

        rename_map = {
            old: f"{label} {new}"
            for old, new in metric_rename.items()
            if old in run_df.columns
        }
        run_df = run_df.rename(columns=rename_map)

        if comparison_df is None:
            comparison_df = run_df
        else:
            comparison_df = comparison_df.merge(run_df, on="transaction", how="outer")

    if comparison_df is None:
        return pd.DataFrame()

    # Add diff columns comparing first and last upload.
    baseline = labels[0]
    latest = labels[-1]

    base_avg = f"{baseline} Avg ResTime in sec"
    latest_avg = f"{latest} Avg ResTime in sec"
    if base_avg in comparison_df.columns and latest_avg in comparison_df.columns:
        comparison_df["Avg ResTime Diff Sec"] = comparison_df[latest_avg] - comparison_df[base_avg]
        comparison_df["Avg ResTime Diff %"] = (comparison_df["Avg ResTime Diff Sec"] / comparison_df[base_avg]) * 100

    base_err = f"{baseline} Error Count"
    latest_err = f"{latest} Error Count"
    if base_err in comparison_df.columns and latest_err in comparison_df.columns:
        comparison_df["Error Count Diff"] = comparison_df[latest_err] - comparison_df[base_err]

    # Put transaction at end because Feature/Scenario/Endpoint columns are the main view.
    ordered = [c for c in comparison_df.columns if c != "transaction"] + ["transaction"]
    return comparison_df[ordered]



def build_comparison_insights_frames(json_paths: List[str | Path], labels: List[str]) -> Dict[str, pd.DataFrame]:
    """Create aggregated frames for Insights when comparing multiple reports."""
    all_apis = []
    all_tx = []
    all_errors = []
    run_info_rows = []

    for path, label in zip(json_paths, labels):
        frames = build_single_report_frames(path)
        region_info = parse_report_metadata(path)
        region = region_info.get("Region", "N/A")

        api_df = frames["APIs"].copy()
        api_df["Run"] = label
        api_df["Region"] = region
        all_apis.append(api_df)

        tx_df = frames["Transactions"].copy()
        tx_df["Run"] = label
        tx_df["Region"] = region
        all_tx.append(tx_df)

        err_df = frames["Errors"].copy()
        err_df["Run"] = label
        err_df["Region"] = region
        all_errors.append(err_df)

        info_row = frames["Run_Info"].iloc[0].to_dict() if "Run_Info" in frames and not frames["Run_Info"].empty else region_info
        info_row["Run"] = label
        run_info_rows.append(info_row)

    combined = {
        "APIs": pd.concat(all_apis, ignore_index=True) if all_apis else pd.DataFrame(),
        "Transactions": pd.concat(all_tx, ignore_index=True) if all_tx else pd.DataFrame(),
        "Errors": pd.concat(all_errors, ignore_index=True) if all_errors else pd.DataFrame(),
        "Run_Info": pd.DataFrame(run_info_rows),
    }
    return combined


def build_comparison_report(json_paths: List[str | Path], labels: List[str], output_excel_path: str | Path) -> None:
    # Comparison mode workbook:
    # Insights uses aggregated results across all uploaded files/regions.
    # Track_Comparison and APIs_Comparison remain side-by-side comparison sheets.
    comparison_insights = build_comparison_insights_frames(json_paths, labels)
    frames = {
        "APIs_Comparison": build_apis_comparison(json_paths, labels)
    }
    track_matrix = build_track_comparison_matrix(json_paths, labels)
    write_excel(
        frames,
        output_excel_path,
        track_matrix=track_matrix,
        insights_frames=comparison_insights,
        comparison_mode=True,
    )


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    subheader_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(color="FFFFFF", bold=True)
    dark_font = Font(color="000000", bold=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A2" if ws.title != "Track_Comparison" else "A3"

    max_header_rows = 1
    for row_idx in range(1, max_header_rows + 1):
        for cell in ws[row_idx]:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            else:
                cell.fill = subheader_fill
                cell.font = dark_font

    for row in ws.iter_rows(min_row=max_header_rows + 1):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, float):
                cell.number_format = "0.00"

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        width = min(max(max_len + 2, 12), 45)
        ws.column_dimensions[get_column_letter(col_idx)].width = width



    # Track_Comparison section styling.
    if ws.title == "Track_Comparison":
        ws.freeze_panes = "A3"
        for row_idx in range(1, ws.max_row + 1):
            first_val = str(ws.cell(row=row_idx, column=1).value or "")
            second_val = str(ws.cell(row=row_idx, column=2).value or "")

            if first_val.startswith("AskAI Tracks"):
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = PatternFill("solid", fgColor="153B50")
                    cell.font = Font(color="FFFFFF", bold=True, size=12)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            elif first_val.startswith("Assets / Assessments / Home / Settings / Support Tracks"):
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = PatternFill("solid", fgColor="1E7D4E")
                    cell.font = Font(color="FFFFFF", bold=True, size=12)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            elif first_val == "Track" and second_val == "Metric":
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = PatternFill("solid", fgColor="D9EAF7")
                    cell.font = Font(color="000000", bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


    # Track_Comparison Total row styling.
    if ws.title == "Track_Comparison":
        for row_idx in range(1, ws.max_row + 1):
            first_val = str(ws.cell(row=row_idx, column=1).value or "")
            if first_val == "Total":
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = PatternFill("solid", fgColor="FFF2CC")
                    cell.font = Font(color="000000", bold=True)
                    cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center", wrap_text=True)

    # Highlight only response-time cells that breach SLA in APIs sheet.
    headers = [cell.value for cell in ws[1]]
    if ws.title == "APIs" and "SLA Sec" in headers:
        sla_col = headers.index("SLA Sec") + 1
        sla_status_col = headers.index("SLA Status") + 1 if "SLA Status" in headers else None
        response_time_columns = [
            "Avg ResTime in sec",
            "Min ResTime in sec",
            "MaxRes Time in sec",
            "90thPercentile Resp Time in Sec",
            "95thPercentile Resp Time in Sec",
            "99thPercentile Resp Time in Sec",
        ]
        target_cols = [headers.index(col_name) + 1 for col_name in response_time_columns if col_name in headers]

        for row in range(2, ws.max_row + 1):
            sla_value = ws.cell(row=row, column=sla_col).value
            try:
                sla_value = float(sla_value)
            except (TypeError, ValueError):
                continue

            for col in target_cols:
                cell = ws.cell(row=row, column=col)
                try:
                    metric_value = float(cell.value)
                except (TypeError, ValueError):
                    continue

                if metric_value >= sla_value:
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")
                    cell.font = Font(color="9C0006", bold=True)

            if sla_status_col:
                status_cell = ws.cell(row=row, column=sla_status_col)
                if status_cell.value == "PASS":
                    status_cell.fill = PatternFill("solid", fgColor="C6EFCE")
                    status_cell.font = Font(color="006100", bold=True)
                elif status_cell.value == "FAIL":
                    status_cell.fill = PatternFill("solid", fgColor="FFC7CE")
                    status_cell.font = Font(color="9C0006", bold=True)


    # APIs: highlight ONLY response-time metric cells that breach SLA.
    # AskAI Feature => 10 sec SLA. Other tracks => 2 sec SLA.
    if ws.title == "APIs":
        headers = [cell.value for cell in ws[1]]
        metric_cols = []
        for col_name in [
            "Avg ResTime in sec",
            "Min ResTime in sec",
            "MaxRes Time in sec",
            "90thPercentile Resp Time in Sec",
            "95thPercentile Resp Time in Sec",
            "99thPercentile Resp Time in Sec",
        ]:
            if col_name in headers:
                metric_cols.append(headers.index(col_name) + 1)

        feature_col = headers.index("Feature") + 1 if "Feature" in headers else (headers.index("Tracks") + 1 if "Tracks" in headers else None)

        if feature_col and metric_cols:
            for row in range(2, ws.max_row + 1):
                feature = str(ws.cell(row=row, column=feature_col).value or "")
                sla_sec = 10 if feature.upper().startswith("ASKAI") else 2

                for col in metric_cols:
                    cell = ws.cell(row=row, column=col)
                    try:
                        value = float(cell.value)
                    except Exception:
                        continue
                    if value >= sla_sec:
                        cell.fill = PatternFill("solid", fgColor="FFC7CE")
                        cell.font = Font(color="9C0006", bold=True)


    # APIs_Comparison: format response-time, error, and diff columns.
    if ws.title == "APIs_Comparison":
        headers = [cell.value for cell in ws[1]]
        diff_cols = []
        for idx, header in enumerate(headers, start=1):
            h = str(header or "").lower()
            if "sec" in h or "error %" in h or "diff %" in h:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=idx).number_format = "0.00"
            if "diff" in h:
                diff_cols.append(idx)

        for col in diff_cols:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                try:
                    value = float(cell.value)
                except Exception:
                    continue
                if value > 0:
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")
                    cell.font = Font(color="9C0006", bold=True)
                elif value < 0:
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                    cell.font = Font(color="006100", bold=True)





def build_insights_sheet(ws, frames: Dict[str, pd.DataFrame]):
    ws.title = "Insights"

    apis_df = frames["APIs"].copy()
    tx_df = frames["Transactions"].copy()

    if not apis_df.empty and "Feature" in apis_df.columns:
        apis_df = apis_df[
            (apis_df["Feature"].astype(str).str.strip().str.lower() != "total")
            & (~apis_df["Feature"].astype(str).str.lower().str.contains("select customer", na=False))
        ].copy()

    total_apis = len(apis_df)
    total_samples = int(pd.to_numeric(apis_df.get("sampleCount", 0), errors="coerce").fillna(0).sum()) if not apis_df.empty else 0
    total_error_count = int(pd.to_numeric(apis_df.get("errorCount", 0), errors="coerce").fillna(0).sum()) if not apis_df.empty else 0

    if not apis_df.empty:
        sla_sec_series = apis_df["Feature"].astype(str).str.upper().str.startswith("ASKAI").map({True: 10, False: 2})
        avg_sec_series = pd.to_numeric(apis_df.get("Avg ResTime in sec", 0), errors="coerce").fillna(0)
        sla_pass = int((avg_sec_series < sla_sec_series).sum())
        sla_fail = int((avg_sec_series >= sla_sec_series).sum())
        avg_resp = round(float(avg_sec_series.mean()), 3)
        p95_avg = round(float(pd.to_numeric(apis_df.get("95thPercentile Resp Time in Sec", 0), errors="coerce").fillna(0).mean()), 3)
    else:
        sla_pass = 0
        sla_fail = 0
        avg_resp = 0
        p95_avg = 0

    sla_pass_pct = round((sla_pass / total_apis) * 100, 2) if total_apis else 0
    sla_fail_pct = round((sla_fail / total_apis) * 100, 2) if total_apis else 0
    sample_error_rate = (total_error_count / total_samples * 100) if total_samples else 0
    health_score = round(max(0, min(100, sla_pass_pct - sample_error_rate)), 2)

    # Clean sheet setup.
    ws.sheet_view.showGridLines = False
    for col in range(1, 14):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["G"].width = 17
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 15
    ws.column_dimensions["J"].width = 15

    # Main title.
    ws.merge_cells("A1:M1")
    ws["A1"] = "CiscoIQ Performance Report App"
    ws["A1"].font = Font(size=20, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="153B50")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # KPI cards.
    cards = [
        ("A3:B5", "Health Score", health_score, "153B50"),
        ("C3:D5", "SLA Pass %", sla_pass_pct, "1E7D4E"),
        ("E3:F5", "SLA Fail %", sla_fail_pct, "A61B1B"),
        ("G3:H5", "Total APIs", total_apis, "31588A"),
        ("I3:J5", "Total Samples", total_samples, "6A4C93"),
        ("K3:L5", "Total Errors", total_error_count, "9A3412"),
    ]
    for cell_range, label, value, color in cards:
        ws.merge_cells(cell_range)
        cell = ws[cell_range.split(":")[0]]
        cell.value = f"{label}\n{value}"
        cell.font = Font(size=13, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Report Context: text-only heading, two-row gap from KPI cards.
    run_info_df = frames.get("Run_Info")
    run_info = run_info_df.iloc[0].to_dict() if run_info_df is not None and not run_info_df.empty else {}

    ws["A8"] = "Report Context"
    ws["A8"].font = Font(size=14, bold=True, color="153B50")
    ws["A8"].alignment = Alignment(horizontal="left", vertical="center")

    context_headers = ["Concurrent Users", "Devices Count", "Date", "Duration", "Region"]
    for idx, header in enumerate(context_headers, start=1):
        header_cell = ws.cell(row=9, column=idx, value=header)
        header_cell.font = Font(bold=True, color="153B50")
        header_cell.alignment = Alignment(horizontal="center", wrap_text=True)
        header_cell.border = Border(bottom=Side(style="thin", color="BFBFBF"))

        value_cell = ws.cell(row=10, column=idx, value=run_info.get(header, "N/A"))
        value_cell.alignment = Alignment(horizontal="center", wrap_text=True)
        value_cell.border = Border(bottom=Side(style="thin", color="BFBFBF"))

    # SLA Breakdown: aligned H:J, no stray columns.
    ws["H8"] = "SLA Breakdown"
    ws["H8"].font = Font(size=14, bold=True, color="153B50")
    ws.merge_cells("H8:J8")

    for col_idx, header in enumerate(["Status", "Count", "Percent"], start=8):
        cell = ws.cell(row=9, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="153B50")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, (status, count, pct) in enumerate(
        [("PASS", sla_pass, sla_pass_pct), ("FAIL", sla_fail, sla_fail_pct), ("TOTAL", sla_pass + sla_fail, 100 if total_apis else 0)],
        start=10,
    ):
        row_values = [status, count, pct]
        for col_offset, value in enumerate(row_values, start=8):
            cell = ws.cell(row=row_idx, column=col_offset, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin", color="BFBFBF"),
                right=Side(style="thin", color="BFBFBF"),
                top=Side(style="thin", color="BFBFBF"),
                bottom=Side(style="thin", color="BFBFBF"),
            )
        if status == "PASS":
            ws.cell(row=row_idx, column=8).font = Font(color="2E7D32", bold=True)
        elif status == "FAIL":
            ws.cell(row=row_idx, column=8).font = Font(color="C62828", bold=True)
        else:
            ws.cell(row=row_idx, column=8).font = Font(color="000000", bold=True)

    # Executive summary and health score explanation.
    ws["A13"] = "Health Score = SLA Pass % minus sample error rate %, bounded between 0 and 100."
    ws["A13"].font = Font(italic=True, color="666666")
    ws.merge_cells("A13:F13")

    ws["A15"] = "Executive Summary"
    ws["A15"].font = Font(size=14, bold=True, color="153B50")
    summary_points = [
        f"SLA result: {sla_pass} APIs passed and {sla_fail} APIs breached SLA.",
        f"Total executed API samples: {total_samples}.",
        f"Average API response time: {avg_resp} sec; average P95 response time: {p95_avg} sec.",
        "AskAI APIs use < 10 sec SLA; Assets, Assessments, Home, Settings and Support APIs use < 2 sec SLA.",
        "Use the ranked tables below to identify the exact APIs behind each chart number.",
    ]
    for row_idx, point in enumerate(summary_points, start=16):
        ws.cell(row=row_idx, column=1, value=f"• {point}")
        ws.cell(row=row_idx, column=1).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)

    # SLA pie chart, with gap below SLA table.
    ws["H15"] = "API SLA Pass vs Fail"
    ws["H15"].font = Font(size=14, bold=True, color="153B50")
    ws.merge_cells("H15:J15")

    pie = PieChart()
    pie.title = None
    labels = Reference(ws, min_col=8, min_row=10, max_row=11)
    data = Reference(ws, min_col=9, min_row=9, max_row=11)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.height = 5
    pie.width = 7
    ws.add_chart(pie, "H17")

    def style_table_header(row_num, start_col, end_col, fill="153B50"):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Top 10 Slow APIs.
    slow_start = 31
    ws["A30"] = "Top 10 Slow APIs"
    ws["A30"].font = Font(size=14, bold=True, color="153B50")
    for idx, header in enumerate(["Rank", "Avg Sec", "Feature", "Scenario", "Endpoint"], start=1):
        ws.cell(row=slow_start, column=idx, value=header)
    style_table_header(slow_start, 1, 5, "31588A")

    top_slow = apis_df.copy()
    if not top_slow.empty:
        top_slow["Avg ResTime in sec"] = pd.to_numeric(top_slow.get("Avg ResTime in sec", 0), errors="coerce").fillna(0)
        top_slow = top_slow.sort_values("Avg ResTime in sec", ascending=False).head(10)

    for rank, (_, row) in enumerate(top_slow.iterrows(), start=1):
        excel_row = slow_start + rank
        ws.cell(row=excel_row, column=1, value=rank)
        ws.cell(row=excel_row, column=2, value=float(row.get("Avg ResTime in sec") or 0))
        ws.cell(row=excel_row, column=3, value=row.get("Feature", ""))
        ws.cell(row=excel_row, column=4, value=row.get("Scenario", ""))
        ws.cell(row=excel_row, column=5, value=row.get("Endpoint", ""))

    slow_chart = BarChart()
    slow_chart.type = "col"
    slow_chart.title = "Top 10 Slow APIs"
    slow_chart.y_axis.title = "Avg Response Time (sec)"
    slow_chart.x_axis.title = "Rank"
    data = Reference(ws, min_col=2, min_row=slow_start, max_row=slow_start + len(top_slow))
    cats = Reference(ws, min_col=1, min_row=slow_start + 1, max_row=slow_start + len(top_slow))
    slow_chart.add_data(data, titles_from_data=True)
    slow_chart.set_categories(cats)
    slow_chart.height = 7
    slow_chart.width = 17
    ws.add_chart(slow_chart, "G31")

    # Top 10 Error APIs.
    err_start = 54
    ws["A53"] = "Top 10 Error APIs"
    ws["A53"].font = Font(size=14, bold=True, color="A61B1B")
    for idx, header in enumerate(["Rank", "Error Count", "Feature", "Scenario", "Endpoint"], start=1):
        ws.cell(row=err_start, column=idx, value=header)
    style_table_header(err_start, 1, 5, "A61B1B")

    top_errors = apis_df.copy()
    if not top_errors.empty:
        top_errors["errorCount"] = pd.to_numeric(top_errors.get("errorCount", 0), errors="coerce").fillna(0)
        top_errors = top_errors[top_errors["errorCount"] > 0].sort_values("errorCount", ascending=False).head(10)

    if top_errors.empty:
        ws.cell(row=err_start + 1, column=1, value=1)
        ws.cell(row=err_start + 1, column=2, value=0)
        ws.cell(row=err_start + 1, column=3, value="No API errors found")
        chart_error_rows = 1
    else:
        for rank, (_, row) in enumerate(top_errors.iterrows(), start=1):
            excel_row = err_start + rank
            ws.cell(row=excel_row, column=1, value=rank)
            ws.cell(row=excel_row, column=2, value=int(row.get("errorCount") or 0))
            ws.cell(row=excel_row, column=3, value=row.get("Feature", ""))
            ws.cell(row=excel_row, column=4, value=row.get("Scenario", ""))
            ws.cell(row=excel_row, column=5, value=row.get("Endpoint", ""))
        chart_error_rows = len(top_errors)

    err_chart = BarChart()
    err_chart.type = "col"
    err_chart.title = "Top 10 Error APIs"
    err_chart.y_axis.title = "Error Count"
    err_chart.x_axis.title = "Rank"
    data = Reference(ws, min_col=2, min_row=err_start, max_row=err_start + chart_error_rows)
    cats = Reference(ws, min_col=1, min_row=err_start + 1, max_row=err_start + chart_error_rows)
    err_chart.add_data(data, titles_from_data=True)
    err_chart.set_categories(cats)
    err_chart.height = 7
    err_chart.width = 17
    ws.add_chart(err_chart, "G54")

    for row in ws.iter_rows(min_row=1, max_row=76, min_col=1, max_col=13):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if isinstance(cell.value, float):
                cell.number_format = "0.00"

    style_sheet(ws)


def add_track_comparison_charts(ws):
    """
    No additional track summary/charts are added.
    The main Track_Comparison table is the source of truth.
    """
    return




def write_track_comparison_sheet(wb: Workbook, track_matrix: List[List[Any]]):
    ws = wb.create_sheet("Track_Comparison")

    for row in track_matrix:
        ws.append(row)

    # HARD cleanup: delete any legacy duplicate row:
    # Track | Metric | <report name> | ...
    # followed by Track | Metric | bucket headers...
    row_idx = 1
    while row_idx < ws.max_row:
        a = ws.cell(row=row_idx, column=1).value
        b = ws.cell(row=row_idx, column=2).value
        c = ws.cell(row=row_idx, column=3).value
        next_a = ws.cell(row=row_idx + 1, column=1).value
        next_b = ws.cell(row=row_idx + 1, column=2).value
        next_c = ws.cell(row=row_idx + 1, column=3).value

        if (
            a == "Track"
            and b == "Metric"
            and c not in ("0-10sec %", "0-2sec %")
            and next_a == "Track"
            and next_b == "Metric"
            and next_c in ("0-10sec %", "0-2sec %")
        ):
            ws.delete_rows(row_idx, 1)
            continue
        row_idx += 1

    # Merge and center section title rows.
    for row_idx in range(1, ws.max_row + 1):
        first_val = str(ws.cell(row=row_idx, column=1).value or "")
        if first_val.startswith("AskAI Tracks") or first_val.startswith("Assets / Assessments / Home / Settings / Support Tracks"):
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=max(ws.max_column, 7))
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    add_track_comparison_charts(ws)
    style_sheet(ws)


def write_excel(frames: Dict[str, pd.DataFrame], output_excel_path: str | Path, track_matrix: List[List[Any]], insights_frames: Dict[str, pd.DataFrame] | None = None, comparison_mode: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    build_insights_sheet(ws, insights_frames if insights_frames is not None else frames)

    # Put Track_Comparison near top after Insights.
    write_track_comparison_sheet(wb, track_matrix)

    sheet_order = ["APIs_Comparison"] if comparison_mode else ["Transactions", "Errors", "APIs", "Comparison"]
    for sheet_name in sheet_order:
        if sheet_name not in frames:
            continue
        ws = wb.create_sheet(sheet_name)
        df = frames[sheet_name]
        if sheet_name == "APIs":
            df = df.drop(
                columns=[c for c in ["SLA Sec", "SLA Rule", "SLA Status", "SLA Breach Sec"] if c in df.columns],
                errors="ignore",
            )
            df = df.rename(columns={"Feature": "Tracks", "Scenario": "Transactions"})
        if sheet_name in ["Transactions", "Errors"]:
            df = df.rename(columns={
                "transaction": "Transaction",
                "sampleCount": "SampleCount",
                "errorCount": "ErrorCount",
                "errorPct": "ErrorPct",
            })
        ws.append(list(df.columns))
        for _, row in df.iterrows():
            ws.append([None if pd.isna(v) else v for v in row.tolist()])
        style_sheet(ws)

    wb.save(output_excel_path)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Generate Excel report from JMeter statistics.json")
    parser.add_argument("json_files", nargs="+", help="One or more JMeter statistics.json files")
    parser.add_argument("--output", default="JMeter_Report.xlsx", help="Path to output .xlsx file")
    args = parser.parse_args()

    labels = [Path(p).stem for p in args.json_files]

    if len(args.json_files) == 1:
        build_report(args.json_files[0], args.output)
    else:
        build_comparison_report(args.json_files, labels, args.output)
